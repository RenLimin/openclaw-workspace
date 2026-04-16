#!/usr/bin/env python3
"""
202602 交付中心月报 v6 - 简化版生成脚本 v2
"""
import openpyxl
from datetime import datetime
import csv
import os
import re
import shutil
from collections import defaultdict, Counter

BASE = os.path.expanduser('~/Downloads/report')
OUTPUT = os.path.join(BASE, '202602交付中心月报-v6.xlsx')

def parse_date(val):
    if not val or str(val).strip() == '':
        return None
    val = str(val).strip()
    for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%Y-%m-%d %H:%M:%S'):
        try:
            return datetime.strptime(val, fmt)
        except ValueError:
            continue
    return None

def read_csv_file(filename):
    filepath = os.path.join(BASE, filename)
    rows = []
    with open(filepath, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            rows.append(row)
    print(f"  {filename}: {len(rows)} rows, {len(reader.fieldnames)} cols")
    return rows

def extract_project_no(name):
    if not name:
        return ''
    m = re.search(r'(SSXM-[\d-]+)', str(name))
    return m.group(1) if m else str(name)

def clear_data_rows(ws, from_row=3):
    for r in range(from_row, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).value = None

def get_legend(ws):
    pm_dept = {}
    team_stat = {}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if row[0]:
            pm_dept[str(row[0]).strip()] = str(row[1]).strip() if row[1] else ''
        if len(row) > 26 and row[25]:
            team_stat[str(row[25]).strip()] = str(row[26]).strip() if row[26] else ''
    return pm_dept, team_stat

def write_basic_row(ws, rn, row_dict, num_cols=40):
    """Write first num_cols columns from CSV row dict to worksheet row."""
    keys = list(row_dict.keys())
    for j in range(min(num_cols, len(keys))):
        val = row_dict[keys[j]]
        if val and str(val).strip():
            ws.cell(row=rn, column=j+1).value = val

STATUS_COLS = {
    '实施未开始': 44, '义务已拆分': 45, '实施进行中': 46, '实施已完成': 47,
    '交付邮件交接中': 48, '交付邮件已归档': 49, '验收文件交接中': 50, '验收文件已归档': 51,
}

def calc_delivery_metrics(ws, rn, row):
    """Calculate delivery metric columns (69-76) for a row."""
    est_del = parse_date(row.get('预估交付完成日期', ''))
    base_del = parse_date(row.get('预算-预估交付完成日期', ''))
    act_start = parse_date(row.get('实际服务/授权开始日期', ''))
    status_val = row.get('状态', '')
    
    # 交付计划准确率"差异" (col 69)
    if est_del and base_del:
        d = (est_del - base_del).days
        ws.cell(row=rn, column=69).value = d
        ws.cell(row=rn, column=70).value = '延后' if d > 0 else ('提前' if d < 0 else '一致')
    elif est_del:
        ws.cell(row=rn, column=69).value = -9999
        ws.cell(row=rn, column=70).value = '当期未填写'
    
    # 按时交付率"差异" (col 73)
    if est_del and act_start:
        d2 = (act_start - est_del).days
        ws.cell(row=rn, column=73).value = d2
        ws.cell(row=rn, column=74).value = '延后' if d2 > 0 else ('提前' if d2 < 0 else '一致')
    elif est_del:
        ws.cell(row=rn, column=73).value = -9999
        ws.cell(row=rn, column=74).value = '当期未填写'
    
    # 是否跨月 - 交付计划 (col 71)
    d1 = ws.cell(row=rn, column=69).value
    if d1 is None or d1 == -9999 or status_val == '履约项交付异常':
        ws.cell(row=rn, column=71).value = '不统计'
    elif est_del and base_del:
        ws.cell(row=rn, column=71).value = '否' if (est_del.year == base_del.year and est_del.month == base_del.month) else '是'
    
    # 是否跨月 - 按时交付 (col 75)
    d2 = ws.cell(row=rn, column=73).value
    if d2 is None or d2 == -9999 or status_val == '履约项交付异常':
        ws.cell(row=rn, column=75).value = '不统计'
    elif est_del and act_start:
        ws.cell(row=rn, column=75).value = '否' if (est_del.year == act_start.year and est_del.month == act_start.month) else '是'
    
    # 考核扣分 - 交付计划 (col 72)
    cm1 = ws.cell(row=rn, column=71).value
    if cm1 in ('一致', '当期未填写', '不统计'):
        ws.cell(row=rn, column=72).value = 0
    elif cm1 == '否' and d1 is not None and abs(d1) < 15:
        ws.cell(row=rn, column=72).value = 0.5
    else:
        ws.cell(row=rn, column=72).value = 1
    
    # 考核扣分 - 按时交付 (col 76)
    cm2 = ws.cell(row=rn, column=75).value
    if cm2 in ('一致', '当期未填写', '不统计'):
        ws.cell(row=rn, column=76).value = 0
    elif cm2 == '否' and d2 is not None and abs(d2) < 15:
        ws.cell(row=rn, column=76).value = 0.5
    else:
        ws.cell(row=rn, column=76).value = 1

def calc_project_status(members):
    """Calculate project-level status for a group of rows."""
    sc = Counter()
    for rn, row in members:
        sc[row.get('状态', '')] += 1
    
    total = len(members)
    delivery = sc.get('交付邮件交接中', 0) + sc.get('交付邮件已归档', 0)
    acceptance = sc.get('验收文件交接中', 0) + sc.get('验收文件已归档', 0)
    implemented = sc.get('实施已完成', 0) + delivery + acceptance
    
    if delivery + acceptance >= total:
        proj_stat = '已结项'
    elif implemented >= total:
        proj_stat = '已完整验收'
    elif implemented > 0:
        proj_stat = '可部分验收'
    else:
        proj_stat = '完全不可验收'
    
    if acceptance > 0:
        lv_stat = '4：正常验收'
    elif delivery > 0:
        lv_stat = '1：正常交付'
    else:
        lv_stat = '2：应交未交'
    
    if acceptance >= total:
        acc_stat = '全部验收'
    elif acceptance > 0:
        acc_stat = '部分验收'
    else:
        acc_stat = '正常验收'
    
    return sc, proj_stat, lv_stat, acc_stat

def main():
    print("=== 202602 交付中心月报 v6 ===\n")
    
    # Copy template
    template_path = os.path.join(BASE, '2026交付月报-模版.xlsx')
    shutil.copy(template_path, OUTPUT)
    print("Template copied.\n")
    
    wb = openpyxl.load_workbook(OUTPUT)
    pm_dept, team_stat = get_legend(wb['图例'])
    print(f"Legend: {len(pm_dept)} PM mappings, {len(team_stat)} team mappings\n")
    
    # Read CSV files
    print("--- CSV Data ---")
    qy_rows = read_csv_file('202602周报-签约项目统计.csv')
    poc_rows = read_csv_file('202602周报-POC&提前实施统计.csv')
    yichang_rows = read_csv_file('202602-签约项目异常处置.csv')
    queshou_rows = read_csv_file('202602确收凭证交接-确收.csv')
    yanshou_rows = read_csv_file('202602确收凭证交接-验收.csv')
    print()
    
    # ========================================================
    # 签约 SHEET
    # ========================================================
    print("--- 签约 ---")
    ws = wb['签约']
    ws.cell(row=1, column=1).value = datetime(2026, 2, 28)
    clear_data_rows(ws, 3)
    
    stat_first = {}
    contract_first = {}
    
    for i, row in enumerate(qy_rows):
        rn = i + 3
        
        # Basic data (all 40 basic columns)
        write_basic_row(ws, rn, row, num_cols=40)
        
        # Col 40: 项目编号
        pn = extract_project_no(row.get('所属项目', ''))
        ws.cell(row=rn, column=40).value = pn
        
        # Col 41: 统计项目编号 (first occurrence)
        if pn and pn not in stat_first:
            stat_first[pn] = rn
            ws.cell(row=rn, column=41).value = pn
        
        # Col 42: 统计合同编号 (first occurrence)
        cn = row.get('销售合同编号', '').strip()
        if cn and cn not in contract_first:
            contract_first[cn] = rn
            ws.cell(row=rn, column=42).value = cn
        
        # Col 43: 合同归档年度
        cd = parse_date(row.get('合同归档日期', ''))
        if cd:
            ws.cell(row=rn, column=43).value = cd.year
    
    # Group by 统计项目编号 for aggregated counts
    sp_groups = defaultdict(list)
    for i, row in enumerate(qy_rows):
        rn = i + 3
        pn = extract_project_no(row.get('所属项目', ''))
        if pn:
            sp_groups[pn].append((rn, row))
    
    for sp, members in sp_groups.items():
        first_rn = members[0][0]
        sc, proj_stat, lv_stat, acc_stat = calc_project_status(members)
        
        # Status counts (cols 44-51)
        for status, col in STATUS_COLS.items():
            ws.cell(row=first_rn, column=col).value = sc.get(status, 0)
        
        # 履约项合计 (col 53)
        ws.cell(row=first_rn, column=53).value = len(members)
        
        # 校验 (col 54)
        sum_s = sum(sc.get(s, 0) for s in STATUS_COLS)
        ws.cell(row=first_rn, column=54).value = len(members) - sum_s
        
        # Project-level status
        ws.cell(row=first_rn, column=55).value = proj_stat
        ws.cell(row=first_rn, column=56).value = lv_stat
        ws.cell(row=first_rn, column=67).value = acc_stat
    
    # Delivery metrics for each row
    for i, row in enumerate(qy_rows):
        rn = i + 3
        calc_delivery_metrics(ws, rn, row)
        
        # 项目经理所属部门 (col 77)
        leader = row.get('负责人', '').strip()
        ws.cell(row=rn, column=77).value = pm_dept.get(leader, '')
        
        # 销售团队-统计 (col 78)
        st = row.get('责任销售所属团队', '').strip()
        ws.cell(row=rn, column=78).value = team_stat.get(st, st)
    
    ws.cell(row=1, column=53).value = len(qy_rows)
    ws.cell(row=1, column=54).value = 0
    print(f"  {len(qy_rows)} rows written")
    
    # ========================================================
    # POC&提前实施 SHEET
    # ========================================================
    print("--- POC&提前实施 ---")
    ws = wb['POC&提前实施']
    ws.cell(row=1, column=1).value = datetime(2026, 2, 28)
    clear_data_rows(ws, 3)
    
    poc_stat_first = {}
    poc_contract_first = {}
    
    for i, row in enumerate(poc_rows):
        rn = i + 3
        write_basic_row(ws, rn, row, num_cols=40)
        
        pn = extract_project_no(row.get('所属项目', ''))
        cn = row.get('销售合同编号', '').strip()
        cd = parse_date(row.get('合同归档日期', ''))
        
        ws.cell(row=rn, column=40).value = pn
        
        if pn and pn not in poc_stat_first:
            poc_stat_first[pn] = rn
            ws.cell(row=rn, column=41).value = pn
        
        if cn and cn not in poc_contract_first:
            poc_contract_first[cn] = rn
            ws.cell(row=rn, column=42).value = cn
        
        if cd:
            ws.cell(row=rn, column=43).value = cd.year
        
        # Status count
        status = row.get('状态', '')
        if status in STATUS_COLS:
            ws.cell(row=rn, column=STATUS_COLS[status]).value = 1
        ws.cell(row=rn, column=53).value = 1
        
        # Project status
        if status in ('实施未开始', '义务已拆分'):
            ws.cell(row=rn, column=55).value = '完全不可验收'
            ws.cell(row=rn, column=56).value = '2：应交未交'
        elif status == '实施进行中':
            ws.cell(row=rn, column=55).value = '完全不可验收'
            ws.cell(row=rn, column=56).value = '1：正常交付'
        elif status == '实施已完成':
            ws.cell(row=rn, column=55).value = '可完整验收'
            ws.cell(row=rn, column=56).value = '1：正常交付'
        ws.cell(row=rn, column=67).value = '正常验收'
        
        # Delivery metrics
        calc_delivery_metrics(ws, rn, row)
        
        # 项目经理所属部门 (col 77)
        leader = row.get('负责人', '').strip()
        ws.cell(row=rn, column=77).value = pm_dept.get(leader, '')
        
        # 销售团队-统计 (col 78)
        st = row.get('责任销售所属团队', '').strip()
        ws.cell(row=rn, column=78).value = team_stat.get(st, st)
        
        # POC-specific (cols 79-83)
        start_date = parse_date(row.get('立项日期', ''))
        if start_date and cd:
            days = max(0, (cd - start_date).days)
            ws.cell(row=rn, column=79).value = days
            if days <= 90:
                ws.cell(row=rn, column=80).value = '3个月以内'
            elif days <= 180:
                ws.cell(row=rn, column=80).value = '3-6个月'
            elif days <= 365:
                ws.cell(row=rn, column=80).value = '6-12个月'
            else:
                ws.cell(row=rn, column=80).value = '超过1年'
        
        ptype = row.get('项目类型(概览)', '')
        ws.cell(row=rn, column=81).value = '已关联' if (cn and ptype == '提前实施') else '未关联'
        ws.cell(row=rn, column=82).value = cd
        ws.cell(row=rn, column=83).value = row.get('所属项目', '')
    
    ws.cell(row=1, column=53).value = len(poc_rows)
    print(f"  {len(poc_rows)} rows written")
    
    # ========================================================
    # 异常项目 SHEET
    # ========================================================
    print("--- 异常项目 ---")
    ws = wb['异常项目']
    clear_data_rows(ws, 2)
    
    for i, row in enumerate(yichang_rows):
        rn = i + 2
        # Write all CSV columns
        keys = list(row.keys())
        for j in range(len(keys)):
            val = row[keys[j]]
            if val and str(val).strip():
                if j + 1 <= ws.max_column:
                    ws.cell(row=rn, column=j+1).value = val
        
        # Col 37: 项目经理团队
        leader = row.get('负责人', '').strip()
        ws.cell(row=rn, column=37).value = pm_dept.get(leader, '')
        
        # Col 38: 项目验收状态
        ws.cell(row=rn, column=38).value = '正常验收'
    
    print(f"  {len(yichang_rows)} rows written")
    
    # ========================================================
    # 确收交接 SHEET
    # ========================================================
    print("--- 确收交接 ---")
    ws = wb['确收交接']
    clear_data_rows(ws, 2)
    
    for i, row in enumerate(queshou_rows):
        rn = i + 2
        ws.cell(row=rn, column=1).value = '2026-02'
        
        col_map = [
            ('标题', 2), ('ID', 3), ('BI履约ID', 4), ('合同编号', 5), ('合同编号1', 5),
            ('客户名称', 6), ('项目经理', 7), ('交接日期', 8),
            ('财务', 9), ('财务是否接收', 10), ('财务反馈', 11),
            ('交付邮件是否跨月', 12), ('PMO', 13), ('PMO备注', 14),
            ('是否修改ones状态', 15), ('是否退回重交', 16), ('备注', 17),
        ]
        for key, col in col_map:
            val = row.get(key, '')
            if val and str(val).strip():
                ws.cell(row=rn, column=col).value = val
        
        cross = row.get('交付邮件是否跨月', '')
        ws.cell(row=rn, column=15).value = '是' if cross == '是' else '否'
        ws.cell(row=rn, column=17).value = '否' if cross == '是' else '是'
    
    print(f"  {len(queshou_rows)} rows written")
    
    # ========================================================
    # 验收交接 SHEET
    # ========================================================
    print("--- 验收交接 ---")
    ws = wb['验收交接']
    clear_data_rows(ws, 2)
    
    for i, row in enumerate(yanshou_rows):
        rn = i + 2
        col_map = [
            ('合同名称', 1), ('标题', 2), ('ID', 3), ('合同编号', 5), ('合同编号1', 5),
            ('客户名称', 9), ('项目经理', 11), ('交接日期', 13),
            ('财务', 17), ('财务是否接收', 18), ('PMO', 22), ('PMO备注', 23),
            ('是否修改ones及OA状态', 24),
        ]
        for key, col in col_map:
            val = row.get(key, '')
            if val and str(val).strip():
                if col <= ws.max_column:
                    ws.cell(row=rn, column=col).value = val
    
    print(f"  {len(yanshou_rows)} rows written")
    
    # ========================================================
    # SAVE
    # ========================================================
    print("\n--- Saving ---")
    wb.save(OUTPUT)
    print(f"Saved: {OUTPUT}")
    
    # Verify
    wb2 = openpyxl.load_workbook(OUTPUT, data_only=True)
    print("\n--- Verification ---")
    for name in ['签约', 'POC&提前实施', '异常项目', '确收交接', '验收交接']:
        ws = wb2[name]
        data_rows = ws.max_row - 2 if ws.max_row > 2 else 0
        print(f"  {name}: {data_rows} data rows")
    
    # Check formula columns
    qy = wb2['签约']
    print(f"\n  签约 Row 3:")
    for c in [31, 32, 34, 35, 37, 38, 40, 41, 43, 53, 55, 56, 67, 69, 70, 73, 74, 77, 78]:
        v = qy.cell(3, c).value
        if v is not None:
            print(f"    Col {c}: {str(v)[:50]}")
    print(f"\n  签约 Row 4:")
    for c in [31, 32, 34, 35, 37, 38, 40, 41, 43, 53, 55, 56, 67, 69, 70, 73, 74, 77, 78]:
        v = qy.cell(4, c).value
        if v is not None:
            print(f"    Col {c}: {str(v)[:50]}")
    wb2.close()
    
    print("\n=== DONE ===")

if __name__ == '__main__':
    main()
