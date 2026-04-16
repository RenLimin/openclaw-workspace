#!/usr/bin/env python3
"""
202602 交付中心月报 - Excel 生成脚本
从 CSV 原始数据填充到模板，保留公式和格式
"""

import openpyxl
import csv
import shutil
import os
from datetime import datetime

# Paths
TEMPLATE = os.path.expanduser('~/Downloads/report/2026交付月报-模版.xlsx')
OUTPUT = os.path.expanduser('~/Downloads/openclaw-skill/202602交付中心月报-v5.xlsx')
DATA_DIR = os.path.expanduser('~/Downloads/report/')

# 1. Copy template to output
shutil.copy2(TEMPLATE, OUTPUT)
print(f"Template copied to: {OUTPUT}")

wb = openpyxl.load_workbook(OUTPUT)
print(f"Sheets: {wb.sheetnames}")

# Helper: read CSV and return (headers, rows)
def read_csv(filename):
    filepath = os.path.join(DATA_DIR, filename)
    with open(filepath, 'r', encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        headers = next(reader)
        rows = list(reader)
    return headers, rows

def parse_date(val):
    """Parse date string to datetime or return as-is"""
    if not val or not val.strip():
        return None
    val = val.strip()
    for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%m/%d/%Y', '%Y-%m-%d %H:%M:%S'):
        try:
            return datetime.strptime(val, fmt)
        except ValueError:
            continue
    return val  # Return as string if not parseable

# ============================================================
# Sheet: 签约 (from 202602周报-签约项目统计.csv)
# ============================================================
print("\n--- Processing 签约 sheet ---")
csv_headers, csv_rows = read_csv('202602周报-签约项目统计.csv')
print(f"  CSV: {len(csv_rows)} data rows, {len(csv_headers)} columns")

# Map template columns to CSV column index
ws = wb['签约']
# Get template headers from row 2
tmpl_headers = {}
for col_idx in range(1, ws.max_column + 1):
    val = ws.cell(row=2, column=col_idx).value
    if val is not None:
        tmpl_headers[val] = col_idx

# Build column mapping: template_col_idx -> csv_col_idx
col_mapping = {}
for tmpl_h, tmpl_col in tmpl_headers.items():
    if tmpl_h in csv_headers:
        csv_col = csv_headers.index(tmpl_h) + 1  # 1-indexed
        col_mapping[tmpl_col] = csv_col

print(f"  Template has {len(tmpl_headers)} named columns")
print(f"  Mapped {len(col_mapping)} columns")

# Clear existing data rows (row 3 onwards) while preserving formulas
# Keep row 1 (date + SUBTOTAL formulas) and row 2 (headers)
for row_idx in range(3, ws.max_row + 1):
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        # Only clear if not a formula
        if cell.value is not None and not str(cell.value).startswith('='):
            cell.value = None

# Write CSV data
for i, csv_row in enumerate(csv_rows):
    row_idx = 3 + i
    for tmpl_col, csv_col in col_mapping.items():
        if csv_col <= len(csv_row):
            val = csv_row[csv_col - 1].strip()
            if val:
                cell = ws.cell(row=row_idx, column=tmpl_col)
                parsed = parse_date(val)
                if isinstance(parsed, datetime):
                    cell.value = parsed
                else:
                    cell.value = val

print(f"  Wrote {len(csv_rows)} rows to 签约 sheet")

# Update date in row 1, column A
ws.cell(row=1, column=1).value = datetime(2026, 2, 28)

# ============================================================
# Sheet: POC&提前实施 (from 202602周报-POC&提前实施统计.csv)
# ============================================================
print("\n--- Processing POC&提前实施 sheet ---")
csv_headers, csv_rows = read_csv('202602周报-POC&提前实施统计.csv')
print(f"  CSV: {len(csv_rows)} data rows, {len(csv_headers)} columns")

ws = wb['POC&提前实施']
tmpl_headers = {}
for col_idx in range(1, ws.max_column + 1):
    val = ws.cell(row=2, column=col_idx).value
    if val is not None:
        tmpl_headers[val] = col_idx

col_mapping = {}
for tmpl_h, tmpl_col in tmpl_headers.items():
    if tmpl_h in csv_headers:
        csv_col = csv_headers.index(tmpl_h) + 1
        col_mapping[tmpl_col] = csv_col

print(f"  Mapped {len(col_mapping)} columns")

# Clear existing data
for row_idx in range(3, ws.max_row + 1):
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        if cell.value is not None and not str(cell.value).startswith('='):
            cell.value = None

# Write data
for i, csv_row in enumerate(csv_rows):
    row_idx = 3 + i
    for tmpl_col, csv_col in col_mapping.items():
        if csv_col <= len(csv_row):
            val = csv_row[csv_col - 1].strip()
            if val:
                cell = ws.cell(row=row_idx, column=tmpl_col)
                parsed = parse_date(val)
                if isinstance(parsed, datetime):
                    cell.value = parsed
                else:
                    cell.value = val

print(f"  Wrote {len(csv_rows)} rows to POC&提前实施 sheet")

# Update date
ws.cell(row=1, column=1).value = datetime(2026, 2, 28)

# ============================================================
# Sheet: 异常项目 (from 202602-签约项目异常处置.csv)
# ============================================================
print("\n--- Processing 异常项目 sheet ---")
csv_headers, csv_rows = read_csv('202602-签约项目异常处置.csv')
print(f"  CSV: {len(csv_rows)} data rows, {len(csv_headers)} columns")

ws = wb['异常项目']
tmpl_headers = {}
for col_idx in range(1, ws.max_column + 1):
    val = ws.cell(row=1, column=col_idx).value
    if val is not None:
        tmpl_headers[val] = col_idx

col_mapping = {}
for tmpl_h, tmpl_col in tmpl_headers.items():
    if tmpl_h in csv_headers:
        csv_col = csv_headers.index(tmpl_h) + 1
        col_mapping[tmpl_col] = csv_col

print(f"  Template headers: {list(tmpl_headers.keys())[:10]}")
print(f"  Mapped {len(col_mapping)} columns")

# Clear existing data (row 2 onwards; header is row 1)
for row_idx in range(2, ws.max_row + 1):
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        if cell.value is not None and not str(cell.value).startswith('='):
            cell.value = None

# Write data
for i, csv_row in enumerate(csv_rows):
    row_idx = 2 + i
    for tmpl_col, csv_col in col_mapping.items():
        if csv_col <= len(csv_row):
            val = csv_row[csv_col - 1].strip()
            if val:
                cell = ws.cell(row=row_idx, column=tmpl_col)
                parsed = parse_date(val)
                if isinstance(parsed, datetime):
                    cell.value = parsed
                else:
                    cell.value = val

print(f"  Wrote {len(csv_rows)} rows to 异常项目 sheet")

# ============================================================
# Sheet: 确收交接 (from 202602确收凭证交接-确收.csv)
# ============================================================
print("\n--- Processing 确收交接 sheet ---")
csv_headers, csv_rows = read_csv('202602确收凭证交接-确收.csv')
print(f"  CSV: {len(csv_rows)} data rows, {len(csv_headers)} columns")

ws = wb['确收交接']
tmpl_headers = {}
for col_idx in range(1, ws.max_column + 1):
    val = ws.cell(row=1, column=col_idx).value
    if val is not None:
        tmpl_headers[val] = col_idx

col_mapping = {}
for tmpl_h, tmpl_col in tmpl_headers.items():
    if tmpl_h in csv_headers:
        csv_col = csv_headers.index(tmpl_h) + 1
        col_mapping[tmpl_col] = csv_col

print(f"  Template headers: {list(tmpl_headers.keys())}")
print(f"  CSV headers: {csv_headers[:10]}...")
print(f"  Mapped {len(col_mapping)} columns")

# Clear existing data (header is row 1, data starts row 2)
for row_idx in range(2, ws.max_row + 1):
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        if cell.value is not None:
            cell.value = None

# Write data
for i, csv_row in enumerate(csv_rows):
    row_idx = 2 + i
    for tmpl_col, csv_col in col_mapping.items():
        if csv_col <= len(csv_row):
            val = csv_row[csv_col - 1].strip()
            if val:
                cell = ws.cell(row=row_idx, column=tmpl_col)
                parsed = parse_date(val)
                if isinstance(parsed, datetime):
                    cell.value = parsed
                else:
                    cell.value = val

print(f"  Wrote {len(csv_rows)} rows to 确收交接 sheet")

# ============================================================
# Sheet: 验收交接 (from 202602确收凭证交接-验收.csv)
# ============================================================
print("\n--- Processing 验收交接 sheet ---")
csv_headers, csv_rows = read_csv('202602确收凭证交接-验收.csv')
print(f"  CSV: {len(csv_rows)} data rows, {len(csv_headers)} columns")

ws = wb['验收交接']
tmpl_headers = {}
for col_idx in range(1, ws.max_column + 1):
    val = ws.cell(row=1, column=col_idx).value
    if val is not None:
        tmpl_headers[val] = col_idx

col_mapping = {}
for tmpl_h, tmpl_col in tmpl_headers.items():
    if tmpl_h in csv_headers:
        csv_col = csv_headers.index(tmpl_h) + 1
        col_mapping[tmpl_col] = csv_col

print(f"  Template headers: {list(tmpl_headers.keys())}")
print(f"  Mapped {len(col_mapping)} columns")

# Clear existing data
for row_idx in range(2, ws.max_row + 1):
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        if cell.value is not None:
            cell.value = None

# Write data
for i, csv_row in enumerate(csv_rows):
    row_idx = 2 + i
    for tmpl_col, csv_col in col_mapping.items():
        if csv_col <= len(csv_row):
            val = csv_row[csv_col - 1].strip()
            if val:
                cell = ws.cell(row=row_idx, column=tmpl_col)
                parsed = parse_date(val)
                if isinstance(parsed, datetime):
                    cell.value = parsed
                else:
                    cell.value = val

print(f"  Wrote {len(csv_rows)} rows to 验收交接 sheet")

# ============================================================
# Save
# ============================================================
wb.save(OUTPUT)
print(f"\nSaved to: {OUTPUT}")

# Verify
wb2 = openpyxl.load_workbook(OUTPUT, data_only=False)
print("\n=== Verification ===")
for s in wb2.sheetnames:
    ws = wb2[s]
    print(f"  {s}: rows={ws.max_row}, cols={ws.max_column}")

print("\nDone!")
