#!/usr/bin/env python3
"""
能力 3: Excel 解析测试
测试 openpyxl、formulas、数据透视表
"""
import os
import sys

def test_openpyxl_basic():
    """测试 3.1: openpyxl 基础解析"""
    print("📊 测试 openpyxl 基础解析...")
    try:
        import openpyxl
        from openpyxl import Workbook
        
        # Create test workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "TestSheet"
        ws["A1"] = "Name"
        ws["B1"] = "Value"
        ws["A2"] = "Item1"
        ws["B2"] = 100
        ws["A3"] = "Item2"
        ws["B3"] = 200
        ws["A4"] = "Total"
        ws["B4"] = "=SUM(B2:B3)"
        
        test_file = "/tmp/test-excel.xlsx"
        wb.save(test_file)
        
        # Re-open and parse
        wb2 = openpyxl.load_workbook(test_file)
        ws2 = wb2.active
        print(f"  工作表: {ws2.title}")
        print(f"  行数: {ws2.max_row}")
        print(f"  列数: {ws2.max_column}")
        print(f"  B4 公式: {ws2['B4'].value}")
        assert ws2.max_row == 4
        assert "SUM" in ws2['B4'].value
        print("  ✅ openpyxl 基础解析通过")
        return True
    except Exception as e:
        print(f"  ❌ openpyxl 失败: {e}")
        return False

def test_formula_extraction():
    """测试 3.2: 公式提取"""
    print("\n🔢 测试公式提取...")
    try:
        import openpyxl
        
        wb = openpyxl.load_workbook("/tmp/test-excel.xlsx")
        ws = wb.active
        
        formulas = []
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas.append({"cell": cell.coordinate, "formula": cell.value})
        
        print(f"  找到 {len(formulas)} 个公式")
        for f in formulas:
            print(f"    {f['cell']}: {f['formula']}")
        assert len(formulas) == 1
        print("  ✅ 公式提取通过")
        return True
    except Exception as e:
        print(f"  ❌ 公式提取失败: {e}")
        return False

def test_formula_execution():
    """测试 3.3: 公式执行"""
    print("\n⚡ 测试公式执行...")
    try:
        import formulas
        
        # Create a simple formula
        func = formulas.Parser().ast("=1+2+3")[1].compile()
        result = func()
        print(f"  1+2+3 = {result}")
        assert result == 6
        print("  ✅ 公式执行通过")
        return True
    except Exception as e:
        print(f"  ❌ 公式执行失败: {e}")
        return False

def test_pivot_table():
    """测试 3.4: 数据透视表解析"""
    print("\n📈 测试数据透视表...")
    try:
        import pandas as pd
        import openpyxl
        
        # Create test data
        data = {
            "Category": ["A", "A", "B", "B", "C"],
            "Value": [10, 20, 30, 40, 50]
        }
        df = pd.DataFrame(data)
        pivot = pd.pivot_table(df, index="Category", values="Value", aggfunc="sum")
        
        print(f"  透视表结果:\n{pivot}")
        assert pivot.loc["A", "Value"] == 30
        assert pivot.loc["C", "Value"] == 50
        print("  ✅ 数据透视表通过")
        return True
    except Exception as e:
        print(f"  ❌ 数据透视表失败: {e}")
        return False

def test_merged_cells():
    """测试 3.5: 合并单元格/样式"""
    print("\n🔲 测试合并单元格和样式...")
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        
        wb = openpyxl.load_workbook("/tmp/test-excel.xlsx")
        ws = wb.active
        
        # Add merged cells
        ws.merge_cells("A5:B5")
        ws["A5"] = "Merged Cell"
        
        # Add styling
        ws["A1"].font = Font(bold=True, color="FF0000")
        ws["A1"].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        test_file2 = "/tmp/test-excel-merged.xlsx"
        wb.save(test_file2)
        
        # Re-open and verify
        wb2 = openpyxl.load_workbook(test_file2)
        ws2 = wb2.active
        print(f"  合并单元格: {list(ws2.merged_cells.ranges)}")
        print(f"  A5 值: {ws2['A5'].value}")
        print(f"  A1 字体: {ws2['A1'].font}")
        assert "A5:B5" in str(ws2.merged_cells.ranges)
        print("  ✅ 合并单元格/样式通过")
        return True
    except Exception as e:
        print(f"  ❌ 合并单元格失败: {e}")
        return False

def main():
    print("=" * 50)
    print("能力 3: Excel 解析测试")
    print("=" * 50)
    
    results = {}
    results["test_3_1_openpyxl"] = "✅ 通过" if test_openpyxl_basic() else "❌ 失败"
    results["test_3_2_formula_extract"] = "✅ 通过" if test_formula_extraction() else "❌ 失败"
    results["test_3_3_formula_exec"] = "✅ 通过" if test_formula_execution() else "❌ 失败"
    results["test_3_4_pivot"] = "✅ 通过" if test_pivot_table() else "❌ 失败"
    results["test_3_5_merged"] = "✅ 通过" if test_merged_cells() else "❌ 失败"
    
    print("\n" + "=" * 50)
    print("测试结果汇总:")
    for k, v in results.items():
        print(f"  {k}: {v}")
    passed = sum(1 for v in results.values() if "✅" in v)
    print(f"  通过率: {passed}/{len(results)}")
    print("=" * 50)
    
    return passed == len(results)

if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)
