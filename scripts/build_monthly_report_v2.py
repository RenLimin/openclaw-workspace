#!/usr/bin/env python3
"""
202602 交付中心月报 - Excel 生成脚本 v2
修复确收交接和验收交接的列映射
"""

import openpyxl
import csv
import shutil
import os
from datetime import datetime

# Paths
TEMPLATE = os.path.expanduser('~/Downloads/report/2026交付月报-模版.xlsx')
OUTPUT = os.path.expanduser('~/Downloads/openclaw-skill/202602交付中心月报-v5.xlsx')
DATA_DIR = os.path.expanduser('~/Downloads/report/')

REPORT_MONTH = "2026年2月"

# 1. Copy template to output
shutil.copy2(TEMPLATE, OUTPUT)
print(f"Template copied to: {OUTPUT}")

wb = openpyxl.load_workbook(OUTPUT)

# Helper: read CSV
def read_csv(filename):
    filepath = os.path.join(DATA_DIR, filename)
    with open(filepath, 'r', encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        headers = next(reader)
        rows = list(reader)
    return headers, rows

def parse_date(val):
    if not val or not val.strip():
        return None
    val = val.strip()
    for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%m/%d/%Y', '%Y-%m-%d %H:%M:%S'):
        try:
            return datetime.strptime(val, fmt)
        except ValueError:
            continue
    return val

# Column alias mapping: template_name -> possible_csv_names
COLUMN_ALIASES = {
    '财务接收人': ['财务接收人', '财务'],
    'PMO反馈': ['PMO反馈', 'PMO'],
    '月份': ['月份'],  # Not in CSV, will be filled manually
    '项目经理所属区域': ['项目经理所属区域', '销售部门'],
}

def find_csv_col(csv_headers, tmpl_name):
    """Find CSV column index for a template column name, using aliases"""
    # Direct match first
    if tmpl_name in csv_headers:
        return csv_headers.index(tmpl_name) + 1
    # Try aliases
    aliases = COLUMN_ALIASES.get(tmpl_name, [tmpl_name])
    for alias in aliases:
        if alias in csv_headers:
            return csv_headers.index(alias) + 1
    return None

def write_csv_to_sheet(ws, csv_headers, csv_rows, header_row=2, data_start_row=3, fill_month=False):
    """Write CSV data to sheet, mapping by header name"""
    # Get template headers
    tmpl_headers = {}
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col_idx).value
        if val is not None:
            tmpl_headers[col_idx] = str(val)

    # Build column mapping
    col_mapping = {}
    for tmpl_col, tmpl_h in tmpl_headers.items():
        csv_col = find_csv_col(csv_headers, tmpl_h)
        if csv_col is not None:
            col_mapping[tmpl_col] = csv_col

    print(f"  Mapped {len(col_mapping)} of {len(tmpl_headers)} template columns")
    print(f"  Unmapped: {[h for c, h in tmpl_headers.items() if c not in col_mapping]}")

    # Clear existing data rows
    for row_idx in range(data_start_row, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None and not str(cell.value).startswith('='):
                cell.value = None

    # Write data
    for i, csv_row in enumerate(csv_rows):
        row_idx = data_start_row + i
        for tmpl_col, csv_col in col_mapping.items():
            if csv_col <= len(csv_row):
                val = csv_row[csv_col - 1].strip()
                if val:
                    cell = ws.cell(row=row_idx, column=tmpl_col)
                    parsed = parse_date(val)
                    if isinstance(parsed, datetime):
                        cell.value = parsed
                    else:
                        cell.value = val

        # Fill month column if needed
        if fill_month:
            month_col = find_csv_col(csv_headers, '月份')
            if month_col is None:
                # Find "月份" column in template
                for tmpl_col, tmpl_h in tmpl_headers.items():
                    if tmpl_h == '月份':
                        ws.cell(row=row_idx, column=tmpl_col).value = REPORT_MONTH
                        break

    return len(csv_rows)

# ============================================================
# Sheet: 签约
# ============================================================
print("\n--- 签约 ---")
csv_headers, csv_rows = read_csv('202602周报-签约项目统计.csv')
ws = wb['签约']
count = write_csv_to_sheet(ws, csv_headers, csv_rows, header_row=2, data_start_row=3)
ws.cell(row=1, column=1).value = datetime(2026, 2, 28)
print(f"  Wrote {count} rows")

# ============================================================
# Sheet: POC&提前实施
# ============================================================
print("\n--- POC&提前实施 ---")
csv_headers, csv_rows = read_csv('202602周报-POC&提前实施统计.csv')
ws = wb['POC&提前实施']
count = write_csv_to_sheet(ws, csv_headers, csv_rows, header_row=2, data_start_row=3)
ws.cell(row=1, column=1).value = datetime(2026, 2, 28)
print(f"  Wrote {count} rows")

# ============================================================
# Sheet: 异常项目
# ============================================================
print("\n--- 异常项目 ---")
csv_headers, csv_rows = read_csv('202602-签约项目异常处置.csv')
ws = wb['异常项目']
count = write_csv_to_sheet(ws, csv_headers, csv_rows, header_row=1, data_start_row=2)
print(f"  Wrote {count} rows")

# ============================================================
# Sheet: 确收交接
# ============================================================
print("\n--- 确收交接 ---")
csv_headers, csv_rows = read_csv('202602确收凭证交接-确收.csv')
ws = wb['确收交接']
count = write_csv_to_sheet(ws, csv_headers, csv_rows, header_row=1, data_start_row=2, fill_month=True)
print(f"  Wrote {count} rows")

# ============================================================
# Sheet: 验收交接
# ============================================================
print("\n--- 验收交接 ---")
csv_headers, csv_rows = read_csv('202602确收凭证交接-验收.csv')
ws = wb['验收交接']
count = write_csv_to_sheet(ws, csv_headers, csv_rows, header_row=1, data_start_row=2, fill_month=True)
print(f"  Wrote {count} rows")

# ============================================================
# Save
# ============================================================
wb.save(OUTPUT)
print(f"\nSaved: {OUTPUT}")

# Verify
wb2 = openpyxl.load_workbook(OUTPUT, data_only=False)
print("\n=== Verification ===")
for s in wb2.sheetnames:
    ws = wb2[s]
    data_rows = ws.max_row - 1 if s in ['签约', 'POC&提前实施'] else ws.max_row - 1 if s in ['异常项目', '确收交接', '验收交接'] else ws.max_row
    print(f"  {s}: rows={ws.max_row}, cols={ws.max_column}")

print("\nDone!")
