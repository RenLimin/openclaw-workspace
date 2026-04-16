#!/usr/bin/env python3
"""
202602 交付中心月报 - Excel 生成脚本 v3 (最终版)
修复所有列映射问题
"""

import openpyxl
import csv
import shutil
import os
from datetime import datetime

TEMPLATE = os.path.expanduser('~/Downloads/report/2026交付月报-模版.xlsx')
OUTPUT = os.path.expanduser('~/Downloads/openclaw-skill/202602交付中心月报-v5.xlsx')
DATA_DIR = os.path.expanduser('~/Downloads/report/')
REPORT_MONTH = "2026年2月"

shutil.copy2(TEMPLATE, OUTPUT)
wb = openpyxl.load_workbook(OUTPUT)

def read_csv(filename):
    filepath = os.path.join(DATA_DIR, filename)
    with open(filepath, 'r', encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        headers = next(reader)
        # Normalize headers: strip whitespace and newlines
        headers = [h.strip().replace('\n', '') for h in headers]
        rows = list(reader)
    return headers, rows

def parse_date(val):
    if not val or not val.strip():
        return None
    val = val.strip()
    for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%m/%d/%Y', '%Y-%m-%d %H:%M:%S'):
        try:
            return datetime.strptime(val, fmt)
        except ValueError:
            continue
    return val

# Column alias mapping: template_name -> possible_csv_names
COLUMN_ALIASES = {
    '财务接收人': ['财务接收人', '财务'],
    'PMO反馈': ['PMO反馈', 'PMO'],
    '月份': ['月份'],
    '项目经理所属区域': ['项目经理所属区域', '销售部门'],
    '是否接收': ['是否接收', '财务是否接收'],
    '是否合格': ['是否合格', '财务是否接收'],  # Fallback: use 财务是否接收
    '跨月交接': ['跨月交接', '交付邮件是否跨月'],
    '合同编号': ['合同编号', '合同编号1'],
}

def find_csv_col(csv_headers, tmpl_name):
    if tmpl_name in csv_headers:
        return csv_headers.index(tmpl_name) + 1
    aliases = COLUMN_ALIASES.get(tmpl_name, [tmpl_name])
    for alias in aliases:
        if alias in csv_headers:
            return csv_headers.index(alias) + 1
    return None

def write_csv_to_sheet(ws, csv_headers, csv_rows, header_row=2, data_start_row=3, 
                        fill_month=False, extra_fill=None):
    tmpl_headers = {}
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col_idx).value
        if val is not None:
            tmpl_headers[col_idx] = str(val).strip()

    col_mapping = {}
    for tmpl_col, tmpl_h in tmpl_headers.items():
        csv_col = find_csv_col(csv_headers, tmpl_h)
        if csv_col is not None:
            col_mapping[tmpl_col] = csv_col

    unmapped = [h for c, h in tmpl_headers.items() if c not in col_mapping]
    print(f"  Mapped {len(col_mapping)} of {len(tmpl_headers)} columns")
    if unmapped:
        print(f"  Unmapped: {unmapped}")

    # Clear existing data
    for row_idx in range(data_start_row, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None and not str(cell.value).startswith('='):
                cell.value = None

    # Find month column in template
    month_tmpl_col = None
    if fill_month:
        for tmpl_col, tmpl_h in tmpl_headers.items():
            if tmpl_h == '月份':
                month_tmpl_col = tmpl_col
                break

    # Write data
    for i, csv_row in enumerate(csv_rows):
        row_idx = data_start_row + i
        for tmpl_col, csv_col in col_mapping.items():
            if csv_col <= len(csv_row):
                val = csv_row[csv_col - 1].strip()
                if val:
                    cell = ws.cell(row=row_idx, column=tmpl_col)
                    parsed = parse_date(val)
                    if isinstance(parsed, datetime):
                        cell.value = parsed
                    else:
                        cell.value = val

        # Fill month
        if month_tmpl_col:
            ws.cell(row=row_idx, column=month_tmpl_col).value = REPORT_MONTH

        # Extra fill logic
        if extra_fill:
            extra_fill(ws, row_idx, csv_row, csv_headers)

    return len(csv_rows)

# ============================================================
# Sheet: 签约
# ============================================================
print("--- 签约 ---")
csv_headers, csv_rows = read_csv('202602周报-签约项目统计.csv')
ws = wb['签约']
write_csv_to_sheet(ws, csv_headers, csv_rows, header_row=2, data_start_row=3)
ws.cell(row=1, column=1).value = datetime(2026, 2, 28)
print(f"  Wrote {len(csv_rows)} rows")

# ============================================================
# Sheet: POC&提前实施
# ============================================================
print("\n--- POC&提前实施 ---")
csv_headers, csv_rows = read_csv('202602周报-POC&提前实施统计.csv')
ws = wb['POC&提前实施']
write_csv_to_sheet(ws, csv_headers, csv_rows, header_row=2, data_start_row=3)
ws.cell(row=1, column=1).value = datetime(2026, 2, 28)
print(f"  Wrote {len(csv_rows)} rows")

# ============================================================
# Sheet: 异常项目
# ============================================================
print("\n--- 异常项目 ---")
csv_headers, csv_rows = read_csv('202602-签约项目异常处置.csv')
ws = wb['异常项目']
write_csv_to_sheet(ws, csv_headers, csv_rows, header_row=1, data_start_row=2)
print(f"  Wrote {len(csv_rows)} rows")

# ============================================================
# Sheet: 确收交接
# ============================================================
print("\n--- 确收交接 ---")
csv_headers, csv_rows = read_csv('202602确收凭证交接-确收.csv')

# Extra fill: 跨月交接原因
def extra_fill_queshou(ws, row_idx, csv_row, csv_headers):
    # 跨月交接原因 - from 备注 column
    idx = csv_headers.index('备注') if '备注' in csv_headers else None
    if idx and idx < len(csv_row):
        val = csv_row[idx].strip()
        if val:
            # Find 跨月交接原因 column
            for c in range(1, ws.max_column + 1):
                if ws.cell(row=1, column=c).value == '跨月交接原因':
                    ws.cell(row=row_idx, column=c).value = val
                    break

ws = wb['确收交接']
write_csv_to_sheet(ws, csv_headers, csv_rows, header_row=1, data_start_row=2, 
                   fill_month=True, extra_fill=extra_fill_queshou)
print(f"  Wrote {len(csv_rows)} rows")

# ============================================================
# Sheet: 验收交接
# ============================================================
print("\n--- 验收交接 ---")
csv_headers, csv_rows = read_csv('202602确收凭证交接-验收.csv')

ws = wb['验收交接']
write_csv_to_sheet(ws, csv_headers, csv_rows, header_row=1, data_start_row=2,
                   fill_month=True)
print(f"  Wrote {len(csv_rows)} rows")

# ============================================================
# Save and verify
# ============================================================
wb.save(OUTPUT)
print(f"\nSaved: {OUTPUT}")

# Final verification
wb2 = openpyxl.load_workbook(OUTPUT, data_only=False)
print("\n=== Final Verification ===")
checks = {
    '签约': (13478, 82),
    'POC&提前实施': (2390, 84),
    '异常项目': (297, 38),
    '确收交接': (223, 17),
    '验收交接': (90, 14),
}
all_ok = True
for sheet, (expected_rows, expected_cols) in checks.items():
    ws = wb2[sheet]
    # data rows = max_row - header_rows
    if sheet in ['签约', 'POC&提前实施']:
        actual_data = ws.max_row - 2
    else:
        actual_data = ws.max_row - 1
    
    status = "✓" if actual_data == expected_rows else "✗"
    if actual_data != expected_rows:
        all_ok = False
    print(f"  {status} {sheet}: {actual_data} data rows (expected {expected_rows}), {ws.max_column} cols")

# Spot check some data
print("\n=== Spot Checks ===")
ws = wb2['确收交接']
print(f"确收交接 row 2: 月份={ws.cell(2,1).value}, 标题={ws.cell(2,2).value}, 跨月交接={ws.cell(2,15).value}")

ws = wb2['验收交接']
print(f"验收交接 row 2: 月份={ws.cell(2,1).value}, 是否接收={ws.cell(2,10).value}, 合同编号={ws.cell(2,5).value}")

ws = wb2['签约']
print(f"签约 row 3: BI履约ID={ws.cell(3,1).value}, 客户={ws.cell(3,3).value}")

print(f"\nAll checks passed: {all_ok}")
print("Done!")
